"""Integration tests for feedback Excel import and export."""

from io import BytesIO
from importlib import import_module

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


def test_feedback_project_admin_can_open_feedback_settings(client: TestClient):
    from automation_hub.core import auth, db

    conn = db.db_connect(db.get_db_file())
    try:
        conn.execute(
            """
            INSERT OR REPLACE INTO users (
                username, password, role, level, modules_json, status
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                "feedback-admin@snapp.cab",
                auth.hash_password("FeedbackAdmin1!"),
                "project_admin",
                "custom",
                '["feedback_180_admin"]',
                "active",
            ),
        )
        conn.commit()
    finally:
        conn.close()
    login = client.post(
        "/login",
        data={
            "username": "feedback-admin@snapp.cab",
            "password": "FeedbackAdmin1!",
        },
        follow_redirects=False,
    )
    assert login.status_code == 302
    page = client.get("/admin/feedback-deadline")
    assert page.status_code == 200
    assert "Feedback project configuration" in page.text
    settings = client.get("/api/feedback/evaluator-nomination/settings")
    assert settings.status_code == 200
    platform_users = client.get("/api/admin/users")
    assert platform_users.status_code == 403


def test_nomination_template_download(authenticated_client: TestClient):
    response = authenticated_client.get(
        "/api/feedback/evaluator-nomination/template.xlsx"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook["Evaluators"]["A1"].value == "email"
    assert workbook["Evaluators"]["B1"].value == "reason"
    assert "Instructions" in workbook.sheetnames


def test_feedback_excel_settings_are_database_driven(
    authenticated_client: TestClient,
):
    response = authenticated_client.post(
        "/api/feedback/evaluator-nomination/settings",
        json={
            "user_deadline": "",
            "manager_deadline": "",
            "excel_max_rows": 120,
            "excel_max_mb": 3,
            "excel_template_columns": [
                "email",
                "reason",
                "team",
                "vertical",
            ],
        },
    )
    assert response.status_code == 200
    assert response.json()["excel"]["max_rows"] == 120

    template = authenticated_client.get(
        "/api/feedback/evaluator-nomination/template.xlsx"
    )
    workbook = load_workbook(BytesIO(template.content), read_only=True)
    headers = [cell.value for cell in workbook["Evaluators"][1]]
    assert headers == ["email", "reason", "team", "vertical"]


def test_nomination_import(authenticated_client: TestClient, monkeypatch):
    router = import_module("automation_hub.projects.feedback.router")

    monkeypatch.setattr(
        router.legacy,
        "_previous_evaluator_keys",
        lambda username: set(),
    )
    monkeypatch.setattr(
        router.legacy.employee_roster,
        "get_employee_by_email",
        lambda email: {
            "username": "person",
            "email": email,
            "e_full_name": "Test Person",
            "team": "Platform",
            "sub_team": "DevOps",
            "vertical": "Technology",
            "job_title": "Engineer",
        },
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evaluators"
    sheet.append(["email", "reason"])
    sheet.append(["person@snapp.cab", "Same project"])
    output = BytesIO()
    workbook.save(output)

    response = authenticated_client.post(
        "/api/feedback/evaluator-nomination/import",
        files={
            "file": (
                "evaluators.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    evaluator = response.json()["evaluators"][0]
    assert evaluator["email"] == "person@snapp.cab"
    assert evaluator["reason"] == "Same project"
    assert evaluator["sub_team"] == "DevOps"


def test_admin_nomination_export(authenticated_client: TestClient, monkeypatch):
    router = import_module("automation_hub.projects.feedback.router")

    monkeypatch.setattr(
        router.legacy,
        "_load_evaluator_store",
        lambda: {
            "nominations": [
                {
                    "id": "NOM_TEST",
                    "nominator_username": "employee",
                    "manager_username": "manager",
                    "submitted_at": "2026-06-15T10:00:00+00:00",
                    "status": "pending",
                    "evaluators": [
                        {
                            "full_name": "Evaluator",
                            "email": "evaluator@snapp.cab",
                            "team": "Platform",
                            "reason": "Same project",
                            "status": "pending",
                        }
                    ],
                }
            ]
        },
    )
    monkeypatch.setattr(
        router.legacy,
        "_get_user_info",
        lambda username: {
            "full_name": "Employee",
            "email": "employee@snapp.cab",
            "team": "Platform",
        },
    )

    response = authenticated_client.get(
        "/api/feedback/evaluator-nomination/export.xlsx"
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook["Nominations"]
    assert sheet["A1"].value == "Nomination ID"
    assert sheet["A2"].value == "NOM_TEST"
    assert sheet["K2"].value == "evaluator@snapp.cab"


def test_single_nomination_export(authenticated_client: TestClient, monkeypatch):
    router = import_module("automation_hub.projects.feedback.router")
    nomination = {
        "id": "NOM_SINGLE",
        "nominator_username": "employee",
        "manager_username": "manager",
        "status": "pending",
        "evaluators": [
            {
                "full_name": "One Evaluator",
                "email": "one@snapp.cab",
                "reason": "Same project",
            }
        ],
    }
    monkeypatch.setattr(
        router.legacy,
        "_load_evaluator_store",
        lambda: {"nominations": [nomination]},
    )
    monkeypatch.setattr(
        router.legacy,
        "_get_user_info",
        lambda username: {
            "full_name": "Employee",
            "email": "employee@snapp.cab",
        },
    )

    response = authenticated_client.get(
        "/api/feedback/evaluator-nomination/NOM_SINGLE/export.xlsx"
    )
    assert response.status_code == 200
    assert "servexa-feedback-employee.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook["Nominations"]["A2"].value == "NOM_SINGLE"
    assert workbook["Nominations"]["K2"].value == "one@snapp.cab"
