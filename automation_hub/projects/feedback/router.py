"""Standalone Feedback API routes.

This router owns the public Feedback API namespace. Ticketing and project
workflow routes live under /api/ticketing.
"""

from datetime import datetime, timezone
from io import BytesIO
import json

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from automation_hub.core import auth, db
from automation_hub.projects.ticketing import router as legacy

router = APIRouter(prefix="/api/feedback", tags=["feedback"])

DEFAULT_IMPORT_MB = 5
DEFAULT_IMPORT_ROWS = 500
DEFAULT_TEMPLATE_COLUMNS = ["email", "reason"]
ALLOWED_TEMPLATE_COLUMNS = [
    "email",
    "reason",
    "full_name",
    "team",
    "sub_team",
    "vertical",
]


def _require_feedback_config_admin(request: Request):
    user = auth.get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    if user.get("role") == "admin":
        return user
    if user.get("role") == "project_admin" and "feedback_180_admin" in (
        user.get("modules") or []
    ):
        return user
    raise HTTPException(status_code=403, detail="Feedback admin access required")


def _excel_settings():
    defaults = {
        "feedback_excel_max_rows": str(DEFAULT_IMPORT_ROWS),
        "feedback_excel_max_mb": str(DEFAULT_IMPORT_MB),
        "feedback_excel_template_columns": json.dumps(DEFAULT_TEMPLATE_COLUMNS),
    }
    conn = db.db_connect(db.get_db_file())
    try:
        rows = conn.execute(
            "SELECT key, value FROM app_settings WHERE key IN (?, ?, ?)",
            tuple(defaults),
        ).fetchall()
    finally:
        conn.close()
    values = {row["key"]: row["value"] for row in rows}
    try:
        columns = json.loads(
            values.get(
                "feedback_excel_template_columns",
                defaults["feedback_excel_template_columns"],
            )
        )
    except (TypeError, ValueError):
        columns = list(DEFAULT_TEMPLATE_COLUMNS)
    columns = [
        column
        for column in ALLOWED_TEMPLATE_COLUMNS
        if column in columns or column in DEFAULT_TEMPLATE_COLUMNS
    ]
    try:
        max_rows = int(values.get("feedback_excel_max_rows", DEFAULT_IMPORT_ROWS))
    except (TypeError, ValueError):
        max_rows = DEFAULT_IMPORT_ROWS
    try:
        max_mb = int(values.get("feedback_excel_max_mb", DEFAULT_IMPORT_MB))
    except (TypeError, ValueError):
        max_mb = DEFAULT_IMPORT_MB
    return {
        "max_rows": max(1, min(5000, max_rows)),
        "max_mb": max(1, min(50, max_mb)),
        "template_columns": columns,
    }


def _xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(sheet, cell_range: str) -> None:
    fill = PatternFill("solid", fgColor="D9F3EA")
    font = Font(bold=True, color="21413A")
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center")


def _build_nomination_export(nominations) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Nominations"
    headers = [
        "Nomination ID",
        "Nominator",
        "Nominator Email",
        "Nominator Team",
        "Sub-team",
        "Vertical",
        "Manager",
        "Submitted At",
        "Request Status",
        "Evaluator",
        "Evaluator Email",
        "Evaluator Team",
        "Evaluator Sub-team",
        "Evaluator Vertical",
        "Reason",
        "Evaluator Status",
    ]
    sheet.append(headers)
    for nomination in nominations:
        nominator = legacy._get_user_info(nomination.get("nominator_username", ""))
        for evaluator in nomination.get("evaluators", []):
            sheet.append(
                [
                    nomination.get("id", ""),
                    nominator.get("full_name", ""),
                    nominator.get("email") or nomination.get("nominator_username", ""),
                    nominator.get("team", ""),
                    nominator.get("sub_team", ""),
                    nominator.get("vertical", ""),
                    nomination.get("manager_username", ""),
                    nomination.get("submitted_at", ""),
                    nomination.get("status", ""),
                    evaluator.get("full_name", ""),
                    evaluator.get("email", ""),
                    evaluator.get("team", ""),
                    evaluator.get("sub_team", ""),
                    evaluator.get("vertical", ""),
                    evaluator.get("reason", ""),
                    evaluator.get("status", "pending"),
                ]
            )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _style_header(sheet, "A1:P1")
    widths = [20, 26, 32, 24, 22, 22, 28, 25, 18, 26, 32, 24, 22, 22, 48, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    return workbook


@router.get("/evaluator-nomination/template.xlsx")
async def evaluator_nomination_template(request: Request):
    legacy._require_feedback_access(request)
    settings = _excel_settings()
    columns = settings["template_columns"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evaluators"
    sheet.append(columns)
    examples = {
        "email": "person@snapp.cab",
        "reason": "Worked closely on the same project",
        "full_name": "Example Person",
        "team": "Platform",
        "sub_team": "DevOps",
        "vertical": "Technology",
    }
    sheet.append([examples[column] for column in columns])
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 64
    last_column = get_column_letter(len(columns))
    _style_header(sheet, f"A1:{last_column}1")
    table = Table(displayName="EvaluatorImport", ref=f"A1:{last_column}2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    guide = workbook.create_sheet("Instructions")
    guide.append(["Servexa evaluator import template"])
    guide.append(["1", "Keep the email and reason headers unchanged."])
    guide.append(["2", "Use one evaluator per row."])
    guide.append(["3", "Email must exist in the Employee Roster."])
    guide.append(["4", "Reason is required."])
    guide.append(["5", f"Maximum {settings['max_rows']} evaluators per file."])
    guide.append(["6", f"Maximum file size is {settings['max_mb']} MB."])
    guide.column_dimensions["A"].width = 8
    guide.column_dimensions["B"].width = 72
    guide["A1"].font = Font(bold=True, size=15, color="21413A")
    guide.merge_cells("A1:B1")
    return _xlsx_response(workbook, "servexa-evaluator-template.xlsx")


@router.post("/evaluator-nomination/import")
async def import_evaluator_nomination(request: Request, file: UploadFile = File(...)):
    user = legacy._require_feedback_access(request)
    legacy._require_nomination_window_open()
    settings = _excel_settings()
    max_import_bytes = settings["max_mb"] * 1024 * 1024
    filename = str(file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    content = await file.read(max_import_bytes + 1)
    if len(content) > max_import_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Excel file must be {settings['max_mb']} MB or smaller",
        )
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = (
            workbook["Evaluators"]
            if "Evaluators" in workbook.sheetnames
            else workbook.active
        )
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid Excel workbook") from exc
    if not raw_headers:
        raise HTTPException(status_code=400, detail="The workbook is empty")
    headers = [str(value or "").strip().lower() for value in raw_headers]
    if "email" not in headers or "reason" not in headers:
        raise HTTPException(
            status_code=400,
            detail="Use the Servexa template and keep email and reason columns",
        )
    email_index = headers.index("email")
    reason_index = headers.index("reason")

    imported = []
    errors = []
    seen = set()
    historical = legacy._previous_evaluator_keys(user.get("username", ""))
    for excel_row, values in enumerate(rows, start=2):
        email = (
            str(values[email_index] or "").strip().lower()
            if values and len(values) > email_index
            else ""
        )
        reason = (
            str(values[reason_index] or "").strip()
            if values and len(values) > reason_index
            else ""
        )
        if not email and not reason:
            continue
        if len(imported) >= settings["max_rows"]:
            errors.append(
                f"Row {excel_row}: maximum {settings['max_rows']} rows exceeded"
            )
            break
        identity = legacy._identity_key(email)
        if not email or not reason:
            errors.append(f"Row {excel_row}: email and reason are required")
            continue
        if identity in seen:
            errors.append(f"Row {excel_row}: duplicate email in this file")
            continue
        if identity in historical:
            errors.append(f"Row {excel_row}: evaluator was nominated previously")
            continue
        roster_user = legacy.employee_roster.get_employee_by_email(email)
        if not roster_user:
            errors.append(f"Row {excel_row}: {email} was not found in Employee Roster")
            continue
        seen.add(identity)
        imported.append(
            {
                "username": roster_user.get("username", ""),
                "email": roster_user.get("email", ""),
                "full_name": roster_user.get("e_full_name")
                or roster_user.get("full_name", ""),
                "job_title": roster_user.get("job_title", ""),
                "team": roster_user.get("team", ""),
                "sub_team": roster_user.get("sub_team", ""),
                "vertical": roster_user.get("vertical", ""),
                "reason": reason,
                "status": "pending",
            }
        )
    if not imported:
        raise HTTPException(
            status_code=400,
            detail=errors[0] if errors else "No valid evaluator rows were found",
        )
    return JSONResponse({"success": True, "evaluators": imported, "errors": errors})


@router.get("/evaluator-nomination/export.xlsx")
async def export_evaluator_nominations(request: Request):
    user = legacy._require_feedback_access(request)
    current_username = legacy._identity_key(user.get("username", ""))
    can_export_all = legacy._can_review_all_feedback(user)
    nominations = legacy._load_evaluator_store().get("nominations", [])
    if not can_export_all:
        nominations = [
            item
            for item in nominations
            if legacy._identity_key(item.get("manager_username")) == current_username
        ]

    workbook = _build_nomination_export(nominations)
    filename = (
        "servexa-all-feedback-nominations.xlsx"
        if can_export_all
        else "servexa-my-team-feedback-nominations.xlsx"
    )
    return _xlsx_response(workbook, filename)


@router.get("/evaluator-nomination/{nomination_id}/export.xlsx")
async def export_single_evaluator_nomination(nomination_id: str, request: Request):
    user = legacy._require_feedback_access(request)
    nomination = next(
        (
            item
            for item in legacy._load_evaluator_store().get("nominations", [])
            if item.get("id") == nomination_id
        ),
        None,
    )
    if not nomination:
        raise HTTPException(status_code=404, detail="Nomination not found")
    current_username = legacy._identity_key(user.get("username", ""))
    if (
        not legacy._can_review_all_feedback(user)
        and legacy._identity_key(nomination.get("manager_username")) != current_username
    ):
        raise HTTPException(
            status_code=403,
            detail="Only the assigned manager, HRBP, or an admin can export this request",
        )
    nominator = legacy._get_user_info(nomination.get("nominator_username", ""))
    filename_key = legacy._identity_key(
        nominator.get("email") or nomination.get("nominator_username", "")
    )
    workbook = _build_nomination_export([nomination])
    return _xlsx_response(
        workbook,
        f"servexa-feedback-{filename_key or nomination_id.lower()}.xlsx",
    )


@router.get("/evaluator-nomination/settings")
async def evaluator_nomination_settings(request: Request):
    _require_feedback_config_admin(request)
    return JSONResponse(
        {
            "success": True,
            "user": legacy._deadline_state(),
            "manager": legacy._manager_deadline_state(),
            "excel": _excel_settings(),
        }
    )


@router.post("/evaluator-nomination/settings")
async def save_evaluator_nomination_settings(request: Request):
    _require_feedback_config_admin(request)
    payload = await request.json()
    deadlines = {}
    for field, setting_key in (
        ("user_deadline", "feedback_nomination_deadline"),
        ("manager_deadline", "feedback_manager_deadline"),
    ):
        raw_deadline = str(payload.get(field) or "").strip()
        normalized = ""
        if raw_deadline:
            try:
                parsed = datetime.fromisoformat(raw_deadline.replace("Z", "+00:00"))
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                normalized = parsed.astimezone(timezone.utc).isoformat()
            except ValueError as exc:
                raise HTTPException(status_code=400, detail="Invalid deadline") from exc
        deadlines[setting_key] = normalized
    try:
        max_rows = max(
            1,
            min(5000, int(payload.get("excel_max_rows") or DEFAULT_IMPORT_ROWS)),
        )
        max_mb = max(1, min(50, int(payload.get("excel_max_mb") or DEFAULT_IMPORT_MB)))
    except (TypeError, ValueError) as exc:
        raise HTTPException(
            status_code=400, detail="Excel limits must be valid numbers"
        ) from exc
    requested_columns = (
        payload.get("excel_template_columns") or DEFAULT_TEMPLATE_COLUMNS
    )
    template_columns = [
        column
        for column in ALLOWED_TEMPLATE_COLUMNS
        if column in requested_columns or column in DEFAULT_TEMPLATE_COLUMNS
    ]
    deadlines.update(
        {
            "feedback_excel_max_rows": str(max_rows),
            "feedback_excel_max_mb": str(max_mb),
            "feedback_excel_template_columns": json.dumps(template_columns),
        }
    )
    conn = db.db_connect(db.get_db_file())
    try:
        for setting_key, value in deadlines.items():
            conn.execute(
                "INSERT OR REPLACE INTO app_settings (key, value) VALUES (?, ?)",
                (setting_key, value),
            )
        conn.commit()
    finally:
        conn.close()
    return JSONResponse(
        {
            "success": True,
            "user": legacy._deadline_state(),
            "manager": legacy._manager_deadline_state(),
            "excel": _excel_settings(),
        }
    )


router.add_api_route(
    "/evaluator-nomination/meta",
    legacy.evaluator_nomination_meta,
    methods=["GET"],
)
router.add_api_route(
    "/evaluator-nomination/users/search",
    legacy.search_evaluator_users,
    methods=["GET"],
)
router.add_api_route(
    "/evaluator-nomination/users/filters",
    legacy.evaluator_user_filters,
    methods=["GET"],
)
router.add_api_route(
    "/evaluator-nomination/my-nomination",
    legacy.get_my_nomination,
    methods=["GET"],
)
router.add_api_route(
    "/evaluator-nomination/history",
    legacy.get_my_nomination_history,
    methods=["GET"],
)
router.add_api_route(
    "/evaluator-nomination/manager/requests",
    legacy.get_manager_requests,
    methods=["GET"],
)
router.add_api_route(
    "/evaluator-nomination/submit",
    legacy.submit_evaluator_nomination,
    methods=["POST"],
)
router.add_api_route(
    "/evaluator-nomination/{nomination_id}/approve-evaluator",
    legacy.approve_evaluator,
    methods=["POST"],
)
router.add_api_route(
    "/evaluator-nomination/{nomination_id}/reject-evaluator",
    legacy.reject_evaluator,
    methods=["POST"],
)
router.add_api_route(
    "/evaluator-nomination/{nomination_id}/add-evaluator",
    legacy.add_evaluator_as_manager,
    methods=["POST"],
)
router.add_api_route(
    "/evaluator-nomination/{nomination_id}/remove-manager-added-evaluator",
    legacy.remove_manager_added_evaluator,
    methods=["POST"],
)
router.add_api_route(
    "/evaluator-nomination/{nomination_id}/close",
    legacy.close_nomination,
    methods=["POST"],
)
